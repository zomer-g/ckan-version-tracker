"""Admin-curated import of מידע לעם (odata) resources into queryable NEON tables.

odata (odata.org.il) is an EXTERNAL CKAN of PROCESSED data. An admin picks
specific datastore resources and pushes each into a real table in the ``odata``
schema of the append DB, so the /data console can query and JOIN them like any
other source — clearly badged "מידע לעם" (processed, not an original public
source).

This is the deliberate, MANUAL counterpart to ``index_mirror``: the same atomic
CSV→NEON load (streaming, all-text, staging-then-swap), but sourced from odata's
datastore CSV dump over HTTP and triggered by an admin action rather than a
version landing. There is no auto-sync — a re-import replaces the table.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
import re
import tempfile
import time
import unicodedata
import uuid
from urllib.parse import urlsplit, unquote

import httpx

from app.config import settings
from app.services import append_store
from app.services.index_mirror import _iter_batches  # schema-agnostic CSV batcher

logger = logging.getLogger(__name__)

SCHEMA = "odata"
REGISTRY = "_imports"
ODATA_BASE = "https://www.odata.org.il"
_UA = "Mozilla/5.0 (over.org.il odata import)"

# A single datastore cell can be large; match the worker / index_mirror cap.
csv.field_size_limit(10**8)


def _qi(name: str) -> str:
    return append_store._qi(name)


def _qt(table: str, schema: str = SCHEMA) -> str:
    return f"{_qi(schema)}.{_qi(table)}"


def _readonly_role() -> str | None:
    """Username of the public console's least-privilege role (see index_mirror).
    None when unset — then the console runs on the read/write role and needs no
    grant."""
    raw = (settings.append_readonly_database_url or "").strip()
    if not raw:
        return None
    return urlsplit(raw).username or None


def table_name_for(resource_id: str, dataset_name: str | None) -> str:
    """Stable, unique, ASCII table name: ``<dataset-slug>_<rid8>``.

    The odata dataset ``name`` is already an ascii slug (e.g. ``amidar-shivook``);
    the resource-id suffix keeps two resources of the same dataset apart. Clipped
    to Postgres' 63-byte identifier budget."""
    base = re.sub(r"[^a-z0-9_]+", "_", (dataset_name or "").lower()).strip("_")
    rid8 = resource_id.replace("-", "")[:8]
    prefix = (base[:44].rstrip("_") + "_") if base else "odata_"
    return f"{prefix}{rid8}"


async def ensure_schema(conn) -> None:
    """Create the ``odata`` schema and make it readable by the console role.
    Idempotent; the GRANTs are what let the read-only console see the tables."""
    await conn.execute(f"CREATE SCHEMA IF NOT EXISTS {_qi(SCHEMA)}")
    role = _readonly_role()
    if not role:
        return
    r = _qi(role)
    try:
        await conn.execute(f"GRANT USAGE ON SCHEMA {_qi(SCHEMA)} TO {r}")
        await conn.execute(f"GRANT SELECT ON ALL TABLES IN SCHEMA {_qi(SCHEMA)} TO {r}")
        await conn.execute(
            f"ALTER DEFAULT PRIVILEGES IN SCHEMA {_qi(SCHEMA)} "
            f"GRANT SELECT ON TABLES TO {r}")
    except Exception:  # noqa: BLE001 — a missing role must not break the import
        logger.warning("odata: could not grant read access to %r", role, exc_info=True)


async def ensure_registry(conn) -> None:
    """Registry of imported resources: which odata resource → which Neon table."""
    await ensure_schema(conn)
    await conn.execute(f"""
        CREATE TABLE IF NOT EXISTS {_qt(REGISTRY)} (
            resource_id  text PRIMARY KEY,
            table_name   text NOT NULL,
            package_id   text,
            dataset_name text,
            title        text,
            organization text,
            format       text,
            source_url   text,
            rows         bigint,
            columns      integer,
            imported_at  timestamptz NOT NULL DEFAULT now()
        )
    """)
    # Added later: the original file URL, kept alongside the resource-page
    # source_url so provenance links survive (see import flow).
    await conn.execute(
        f"ALTER TABLE {_qt(REGISTRY)} ADD COLUMN IF NOT EXISTS source_file_url text")


async def _record_import(*, resource_id: str, table: str, package_id: str | None,
                         dataset_name: str | None, title: str, organization: str | None,
                         fmt: str, source_url: str, file_url: str | None,
                         rows: int, columns: int) -> None:
    """Upsert the registry row, grant the console role SELECT, drop the catalog
    cache — shared by both the server-side and client-upload import paths."""
    pool = await append_store.get_pool()
    async with pool.acquire() as conn:
        await ensure_registry(conn)
        await conn.execute(f"""
            INSERT INTO {_qt(REGISTRY)}
                (resource_id, table_name, package_id, dataset_name, title,
                 organization, format, source_url, source_file_url, rows,
                 columns, imported_at)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11, now())
            ON CONFLICT (resource_id) DO UPDATE SET
                table_name = EXCLUDED.table_name,
                package_id = EXCLUDED.package_id,
                dataset_name = EXCLUDED.dataset_name,
                title = EXCLUDED.title,
                organization = EXCLUDED.organization,
                format = EXCLUDED.format,
                source_url = EXCLUDED.source_url,
                source_file_url = EXCLUDED.source_file_url,
                rows = EXCLUDED.rows,
                columns = EXCLUDED.columns,
                imported_at = now()
        """, resource_id, table, package_id, dataset_name, title, organization,
             fmt, source_url, file_url, rows, columns)
        role = _readonly_role()
        if role:
            try:
                await conn.execute(f"GRANT SELECT ON {_qt(table)} TO {_qi(role)}")
            except Exception:  # noqa: BLE001
                logger.warning("odata: grant on %s failed", table, exc_info=True)
    from app.services.data_catalog import invalidate_catalog_cache
    invalidate_catalog_cache()


async def import_uploaded(
    *, resource_id: str, fmt: str, dataset_name: str | None, title: str | None,
    organization: str | None, source_url: str | None, file_url: str | None,
    tmp_path: str, filename: str | None = None, progress=None,
) -> dict:
    """Import a file the ADMIN's BROWSER already downloaded and uploaded to us.

    odata's file downloads are behind Cloudflare and 403 datacenter IPs (Render),
    while the admin's browser can fetch them (CORS is open). So the file arrives
    as an upload; we only parse + load it here. Provenance links (source_url =
    resource page, file_url = original file) are stored with the table."""
    if not append_store.is_configured():
        raise RuntimeError("append DB is not configured (APPEND_DATABASE_URL missing)")
    f = infer_format(fmt, filename, file_url)
    if f not in SUPPORTED_FILE_FORMATS:
        raise ValueError(
            f"פורמט לא נתמך: {fmt or '—'}. נתמכים: CSV, XLS, XLSX, ICS/ICAL.")
    table = table_name_for(resource_id, dataset_name)
    columns, batches = await asyncio.to_thread(_open_for_load, tmp_path, f)
    if not columns:
        raise ValueError("לא נמצאו עמודות בקובץ")
    loaded = await _load_rows(table, columns, batches, progress=progress)
    final_title = (title or "").strip() or dataset_name or resource_id
    src = source_url or (
        f"{ODATA_BASE}/dataset/{dataset_name}/resource/{resource_id}"
        if dataset_name else f"{ODATA_BASE}/dataset")
    await _record_import(
        resource_id=resource_id, table=table, package_id=None,
        dataset_name=dataset_name, title=final_title, organization=organization,
        fmt=f, source_url=src, file_url=file_url,
        rows=loaded["rows"], columns=loaded["columns"])
    logger.info("odata import (upload): %s -> odata.%s (%d rows, %s)",
                resource_id, table, loaded["rows"], f)
    return {"resource_id": resource_id, "table": table, "title": final_title,
            "organization": organization, "rows": loaded["rows"],
            "columns": loaded["columns"], "source_url": src, "format": f}


async def _fetch_json(client: httpx.AsyncClient, action: str, **params) -> dict:
    r = await client.get(
        f"{ODATA_BASE}/api/3/action/{action}",
        params=params,
        headers={"User-Agent": _UA},
    )
    r.raise_for_status()
    data = r.json()
    if not data.get("success"):
        raise ValueError(f"odata {action} failed: {data.get('error')}")
    return data["result"]


async def _download_dump(resource_id: str, path: str) -> None:
    """Stream the resource's whole datastore table as CSV to ``path``."""
    url = f"{ODATA_BASE}/datastore/dump/{resource_id}"
    async with httpx.AsyncClient(
        timeout=httpx.Timeout(600.0), follow_redirects=True
    ) as client:
        async with client.stream(
            "GET", url, params={"format": "csv"},
            headers={"User-Agent": _UA},
        ) as resp:
            resp.raise_for_status()
            with open(path, "wb") as fh:
                async for chunk in resp.aiter_bytes(64 * 1024):
                    fh.write(chunk)


# ── file-format parsing (ported from OCAL's ckan.ts import engine) ───────────
# odata resources come as CSV/XLS/XLSX (spreadsheets) or ICS/ICAL (calendars).
# Spreadsheets are downloaded RAW (not via the datastore) because CKAN's
# datastore mangles Hebrew column names into ASCII; parsing the file keeps them.

SPREADSHEET_FORMATS = {"XLS", "XLSX"}
ICAL_FORMATS = {"ICS", "ICAL", "ICA"}
SUPPORTED_FILE_FORMATS = {"CSV"} | SPREADSHEET_FORMATS | ICAL_FORMATS

# File extension → format, for resources odata publishes with a BLANK format.
_EXT_FORMATS = {
    "csv": "CSV", "xlsx": "XLSX", "xlsm": "XLSX", "xls": "XLS",
    "ics": "ICS", "ical": "ICAL", "ica": "ICA",
}


def _ext_format(candidate: str | None) -> str:
    """Supported format implied by a file name or download URL, else ""."""
    s = (candidate or "").strip()
    if not s:
        return ""
    # A URL's extension lives in the path — never in the query string.
    path = urlsplit(s).path if "://" in s else s
    ext = os.path.splitext(unquote(path))[1].lstrip(".").lower()
    return _EXT_FORMATS.get(ext, "")


def infer_format(fmt: str | None, *candidates: str | None) -> str:
    """The resource's real format: its declared ``format`` when that is one we
    support, otherwise the extension of its file name / download URL.

    Some odata resources carry an EMPTY ``format`` — "רשימת כתובות בישראל עם
    קואורדינטות" is one, published as ``כתובות.xlsx`` with ``format: ""`` and no
    datastore. Trusting the declared field alone made those files un-importable
    even though the bytes are a perfectly ordinary spreadsheet."""
    declared = (fmt or "").strip().upper()
    if declared in SUPPORTED_FILE_FORMATS:
        return declared
    for c in candidates:
        found = _ext_format(c)
        if found:
            return found
    return declared

# A cell holds a real header if it is a string carrying a Hebrew or Latin letter.
_LETTER_RE = re.compile("[A-Za-zא-ת]")
# Invisible marks Israeli gov files embed in headers (RTL/LTR/BOM/ZWS/…).
_INVISIBLE_RE = re.compile(
    "[​-‏﻿­  ‪-‮⁠]")


def is_supported_file_format(fmt: str | None, *candidates: str | None) -> bool:
    return infer_format(fmt, *candidates) in SUPPORTED_FILE_FORMATS


def _norm_header(s: str) -> str:
    """NFC-normalise, strip invisible directional marks, collapse whitespace."""
    s = unicodedata.normalize("NFC", s or "")
    s = _INVISIBLE_RE.sub("", s)
    return re.sub(r"\s+", " ", s).strip()


async def _download_file(url: str, path: str) -> None:
    """Stream a resource's raw file to ``path``."""
    async with httpx.AsyncClient(
        timeout=httpx.Timeout(600.0), follow_redirects=True
    ) as client:
        async with client.stream(
            "GET", url, headers={"User-Agent": _UA},
        ) as resp:
            resp.raise_for_status()
            with open(path, "wb") as fh:
                async for chunk in resp.aiter_bytes(64 * 1024):
                    fh.write(chunk)


# How many leading rows may be inspected when looking for the header row, and
# how many when sizing up a sheet. Both are windows on purpose: a streaming read
# must decide from a bounded prefix (see _stream_table).
_HEADER_SCAN_ROWS = 20
_SHEET_PEEK_ROWS = 50


def _pick_header(head: list[list]) -> tuple[list[str], list[int], int]:
    """Find the header row inside a scan window → (columns, kept indices, row).

    Gov files often open with merged title/logo rows, so the header is the first
    row with the most letter-bearing string cells. Returns ``([], [], -1)`` when
    the window holds no such row (a header-less file)."""
    best_i, best_score = 0, -1
    for i, r in enumerate(head):
        score = sum(1 for v in r if isinstance(v, str) and _LETTER_RE.search(v))
        if score > best_score:
            best_score, best_i = score, i
        if best_score >= 5:
            break
    if best_score <= 0:
        return [], [], -1
    raw = [_norm_header(str(v)) if v is not None else "" for v in head[best_i]]
    safe = append_store.safe_column_names(raw)
    keep = [j for j, c in enumerate(safe) if c and c not in ("_id", "_full_text")]
    return [safe[j] for j in keep], keep, best_i


def _project(row, keep: list[int]) -> list:
    """One raw cell row → the kept columns, as text."""
    return [None if j >= len(row) or row[j] is None else str(row[j]) for j in keep]


def _rows_to_table(rows: list[list]) -> tuple[list[str], list[list]]:
    """A matrix of cell values → (safe column names, data rows).

    Ports OCAL's header handling: auto-detect the real header row, strip
    invisible marks, and fall back to synthetic ``col_N`` headers for
    header-less files."""
    if not rows:
        return [], []
    columns, keep, hdr_i = _pick_header(rows[:_HEADER_SCAN_ROWS])
    if hdr_i < 0:
        # Header-less: synthesise col_1..col_n, every row is data.
        ncols = max((len(r) for r in rows), default=0)
        columns = [f"col_{k + 1}" for k in range(ncols)]
        keep = list(range(ncols))
    return columns, [_project(r, keep) for r in rows[hdr_i + 1:]]


def _stream_table(row_iter) -> tuple[list[str], object]:
    """Streaming twin of ``_rows_to_table``: an iterator of raw cell rows →
    (columns, generator of projected rows).

    Only the header-scan window is ever held in memory, so a half-million-row
    sheet costs a few kilobytes here instead of the ~240MB the eager parser
    needed on a 512MB dyno. The one behavioural difference: a header-less file's
    width is measured over the scan window rather than every row, because a
    single forward pass cannot know the widest row in advance."""
    it = iter(row_iter)
    head: list[list] = []
    for r in it:
        head.append(list(r))
        if len(head) >= _HEADER_SCAN_ROWS:
            break
    if not head:
        return [], iter(())
    columns, keep, hdr_i = _pick_header(head)
    if hdr_i < 0:
        ncols = max((len(r) for r in head), default=0)
        columns = [f"col_{k + 1}" for k in range(ncols)]
        keep = list(range(ncols))
    rest = head[hdr_i + 1:]

    def _gen():
        for r in rest:
            yield _project(r, keep)
        for r in it:
            yield _project(r, keep)

    return columns, _gen()


def _widest_sheet(wb):
    """The workbook's data sheet — the one widest over its first rows.

    Peeking a window (rather than reading every sheet whole, as the eager parser
    did) still skips narrow chart/notes tabs without loading anything big."""
    best, best_cols = None, -1
    for ws in wb.worksheets:
        ncols = 0
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            ncols = max(ncols, len(r))
            if i + 1 >= _SHEET_PEEK_ROWS:
                break
        if ncols > best_cols:
            best, best_cols = ws, ncols
    return best


def _stream_xlsx(path: str) -> tuple[list[str], object]:
    """(columns, row generator) over an XLSX, read-only and forward-only.

    The workbook stays open for as long as the generator lives and is closed
    when it is exhausted or thrown away."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _widest_sheet(wb)
        if ws is None:
            wb.close()
            return [], iter(())
        columns, rows = _stream_table(ws.iter_rows(values_only=True))
        if not columns:                       # nothing to read — close now, the
            wb.close()                        # guarded generator never starts
            return [], iter(())
    except BaseException:
        wb.close()
        raise

    def _guarded():
        try:
            yield from rows
        finally:
            wb.close()

    return columns, _guarded()


def _parse_xlsx(path: str) -> tuple[list[str], list[list]]:
    columns, rows = _stream_xlsx(path)
    return columns, list(rows)


def _parse_xls(path: str) -> tuple[list[str], list[list]]:
    import xlrd
    book = xlrd.open_workbook(path)
    best = max(book.sheets(), key=lambda s: s.ncols, default=None)
    if best is None:
        return [], []
    rows = [best.row_values(i) for i in range(best.nrows)]
    return _rows_to_table(rows)


# How much of a CSV is read to decide its encoding. The eager parser could
# examine every byte; a streaming one has to choose from a prefix. 4MB is
# thousands of rows — orders of magnitude more than the header — and the
# alternative (a full pass just to sniff) is the cost this whole function
# exists to avoid.
_CSV_SNIFF_BYTES = 4 * 1024 * 1024


def _csv_encoding(path: str) -> str:
    """utf-8-sig, or windows-1255 for the many Israeli gov CSVs published in it.

    Same rule the eager parser used — no Hebrew after a UTF-8 decode, but bytes
    sitting in the win-1255 Hebrew range — applied to a prefix instead of the
    whole file.
    """
    with open(path, "rb") as fh:
        raw = fh.read(_CSV_SNIFF_BYTES)
    text = raw.decode("utf-8-sig", errors="replace")
    if not re.search(r"[א-ת]", text) and any(0xC0 <= b <= 0xFA for b in raw):
        try:
            raw.decode("windows-1255")
            return "windows-1255"
        except Exception:  # noqa: BLE001 — keep utf-8 when the guess will not decode
            pass
    return "utf-8-sig"


def _stream_csv(path: str) -> tuple[list[str], object]:
    """(columns, row generator) over a CSV, forward-only.

    The streaming twin of _parse_csv_file, and the reason it exists: that one
    held the whole file as bytes, then again as a str, then again as a list of
    millions of small lists — all three alive at once. On גזטיר נכסים (413MB)
    that is several GB, and it OOM-killed a 2GB dyno. Nothing downstream ever
    needed the rows at once: _batches_from_rows is already a generator and the
    loader consumes it in COPY batches.

    The file stays open for as long as the generator lives and is closed when it
    is exhausted or thrown away — same contract as _stream_xlsx.
    """
    fh = open(path, "r", encoding=_csv_encoding(path), errors="replace",
              newline="")
    try:
        columns, rows = _stream_table(csv.reader(fh))
        if not columns:
            fh.close()
            return [], iter(())
    except BaseException:
        fh.close()
        raise

    def _guarded():
        try:
            yield from rows
        finally:
            fh.close()

    return columns, _guarded()


def _parse_csv_file(path: str) -> tuple[list[str], list[list]]:
    """Eager CSV read. Kept for callers that genuinely want the whole table in
    memory; the import path streams instead — see _stream_csv."""
    columns, rows = _stream_csv(path)
    return columns, [list(r) for r in rows]


def _ical_dt(prop) -> str:
    if prop is None:
        return ""
    try:
        dt = getattr(prop, "dt", prop)
        return dt.isoformat() if hasattr(dt, "isoformat") else str(dt)
    except Exception:  # noqa: BLE001
        return str(prop)


def _parse_ical(path: str) -> tuple[list[str], list[list]]:
    from icalendar import Calendar
    with open(path, "rb") as fh:
        cal = Calendar.from_ical(fh.read())
    columns = ["title", "start_time", "end_time", "location", "description",
               "organizer", "participants", "uid", "status"]
    data: list[list] = []
    for comp in cal.walk("VEVENT"):
        def g(k: str) -> str:
            v = comp.get(k)
            return "" if v is None else str(v)
        organizer = g("organizer").replace("mailto:", "").replace("MAILTO:", "")
        att = comp.get("attendee")
        attendees = att if isinstance(att, list) else ([att] if att else [])
        parts: list[str] = []
        for a in attendees:
            cn = None
            try:
                cn = a.params.get("CN")
            except Exception:  # noqa: BLE001
                cn = None
            parts.append(str(cn) if cn else
                         str(a).replace("mailto:", "").replace("MAILTO:", ""))
        data.append([
            g("summary"), _ical_dt(comp.get("dtstart")), _ical_dt(comp.get("dtend")),
            g("location"), g("description"), organizer, ", ".join(parts),
            g("uid"), g("status"),
        ])
    return columns, data


def _parse_file(path: str, fmt: str) -> tuple[list[str], list[list]]:
    """Dispatch a downloaded file to the right parser → (columns, rows)."""
    f = (fmt or "").upper()
    if f in ICAL_FORMATS:
        return _parse_ical(path)
    if f in SPREADSHEET_FORMATS:
        # odata mislabels some resources (an .xlsx served as format "XLS", etc.),
        # so trust the file's magic bytes over the declared format: a ZIP header
        # (PK) is xlsx/xlsm; an OLE2 header is a legacy .xls.
        with open(path, "rb") as fh:
            head = fh.read(8)
        if head[:2] == b"PK":
            return _parse_xlsx(path)
        if head[:4] == b"\xd0\xcf\x11\xe0":
            return _parse_xls(path)
        return _parse_xlsx(path) if f == "XLSX" else _parse_xls(path)
    if f == "CSV":
        return _parse_csv_file(path)
    raise ValueError(f"unsupported format: {fmt}")


def _batches_from_rows(rows: list[list], ncols: int):
    """Yield COPY batches of tuples from parsed rows, bounded by rows AND bytes
    (matches _iter_batches). Rows are normalised to exactly ``ncols`` values."""
    batch: list[tuple] = []
    nbytes = 0
    for r in rows:
        rec = tuple(r[j] if j < len(r) else None for j in range(ncols))
        batch.append(rec)
        nbytes += sum(len(v) for v in rec if isinstance(v, str))
        if len(batch) >= 20_000 or nbytes >= 16 * 1024 * 1024:
            yield batch
            batch, nbytes = [], 0
    if batch:
        yield batch


def _open_for_load(path: str, fmt: str) -> tuple[list[str], object]:
    """A downloaded file → (columns, iterator of COPY batches).

    XLSX streams (the format big gov spreadsheets arrive in — see _stream_xlsx);
    the rest are parsed eagerly, which is fine for the sizes they come in: XLS
    caps at 65k rows and calendars are tiny. Blocking — call in a thread."""
    f = (fmt or "").upper()
    if f == "XLSX":
        columns, rows = _stream_xlsx(path)
        return columns, _batches_from_rows(rows, len(columns))
    if f == "CSV":
        # CSV was in "the rest" and parsed eagerly, on the assumption that only
        # spreadsheets get big. CSV is exactly the format that gets big: גזטיר
        # נכסים is 413MB, and reading it whole OOM-killed a 2GB dyno.
        columns, rows = _stream_csv(path)
        return columns, _batches_from_rows(rows, len(columns))
    # XLS caps at 65k rows and calendars are tiny, so eager is fine for those.
    columns, rows = _parse_file(path, fmt)
    return columns, _batches_from_rows(rows, len(columns))


# ── NEON loaders ─────────────────────────────────────────────────────────────

async def _aiter_batches(batch_iter):
    """Pull COPY batches off a BLOCKING iterator without stalling the loop.

    Parsing is CPU-bound and, for a half-million-row spreadsheet, runs for
    minutes; consuming the generator inline would freeze every other request
    for its whole duration. One thread hop per 20k-row batch is free by
    comparison."""
    it = iter(batch_iter)
    done = object()
    while True:
        batch = await asyncio.to_thread(next, it, done)
        if batch is done:
            return
        yield batch


async def _load_rows(table: str, columns: list[str], batch_iter,
                     progress=None) -> dict:
    """Create ``odata.<table>`` from ``columns`` + an iterable of COPY batches
    (each a list of tuples), replacing any existing table atomically.

    Every column is ``text``; rows land in a staging table and a single
    transaction drops the old table and renames staging into place, so readers
    see the old table or the new one, never a partial load. ``progress``, if
    given, is called with the running row count after each batch."""
    if not columns:
        raise ValueError("no usable columns")
    staging = append_store.clip_ident_bytes(table, 63 - len("__stg")) + "__stg"
    defs = ", ".join(f"{_qi(c)} text" for c in columns)
    pool = await append_store.get_pool()
    async with pool.acquire() as conn:
        await ensure_schema(conn)
        await conn.execute(f"DROP TABLE IF EXISTS {_qt(staging)}")
        await conn.execute(f"CREATE TABLE {_qt(staging)} ({defs})")
        rows = 0
        try:
            async for batch in _aiter_batches(batch_iter):
                if not batch:
                    continue
                await conn.copy_records_to_table(
                    staging, schema_name=SCHEMA, columns=columns, records=batch,
                )
                rows += len(batch)
                if progress is not None:
                    progress(rows)
            async with conn.transaction():
                await conn.execute(f"DROP TABLE IF EXISTS {_qt(table)}")
                await conn.execute(
                    f"ALTER TABLE {_qt(staging)} RENAME TO {_qi(table)}")
        except BaseException:
            try:
                await conn.execute(f"DROP TABLE IF EXISTS {_qt(staging)}")
            except Exception:  # noqa: BLE001 — cleanup is best-effort
                logger.debug("odata: staging cleanup failed for %s", staging,
                             exc_info=True)
            raise
        await conn.execute(f"ANALYZE {_qt(table)}")
    return {"rows": rows, "columns": len(columns)}


async def _load_csv_into(path: str, table: str) -> dict:
    """Load a datastore-dump CSV file into ``odata.<table>`` (streaming)."""
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        header = next(csv.reader(fh), None)
    if not header:
        raise ValueError("dump CSV is empty (no header row)")
    safe = append_store.safe_column_names(header)
    keep = [i for i, c in enumerate(safe) if c and c not in ("_id", "_full_text")]
    columns = [safe[i] for i in keep]
    if not columns:
        raise ValueError("dump CSV has no usable columns")
    return await _load_rows(table, columns, _iter_batches(path, columns, keep))


async def import_resource(resource_id: str) -> dict:
    """Import one odata resource into a queryable ``odata`` table.

    Handles CSV / XLS / XLSX / ICS / ICAL (ported from OCAL), plus any
    datastore-active resource via its CSV dump. Spreadsheets and calendars are
    downloaded RAW and parsed (keeps Hebrew column names). Raises ValueError for
    an unsupported, non-datastore resource."""
    if not append_store.is_configured():
        raise RuntimeError("append DB is not configured (APPEND_DATABASE_URL missing)")

    async with httpx.AsyncClient(
        timeout=httpx.Timeout(60.0), follow_redirects=True
    ) as client:
        res = await _fetch_json(client, "resource_show", id=resource_id)
        pkg: dict = {}
        if res.get("package_id"):
            try:
                pkg = await _fetch_json(client, "package_show", id=res["package_id"])
            except Exception:  # noqa: BLE001 — metadata is nice-to-have
                logger.debug("odata: package_show failed for %s", res.get("package_id"),
                             exc_info=True)

    url = (res.get("url") or "").strip()
    # A blank ``format`` is common on odata — fall back to the file extension.
    fmt = infer_format(res.get("format"), res.get("name"), url)
    datastore_active = bool(res.get("datastore_active"))
    supported = fmt in SUPPORTED_FILE_FORMATS
    if not supported and not datastore_active:
        raise ValueError(
            f"פורמט לא נתמך: {res.get('format') or '—'}. "
            f"נתמכים: CSV, XLS, XLSX, ICS/ICAL, או משאב עם datastore.")

    dataset_name = pkg.get("name")
    title = (pkg.get("title") or "").strip() or (res.get("name") or "").strip() \
        or dataset_name or resource_id
    org = ((pkg.get("organization") or {}) or {}).get("title")
    table = table_name_for(resource_id, dataset_name)
    source_url = (
        f"{ODATA_BASE}/dataset/{dataset_name}/resource/{resource_id}"
        if dataset_name else f"{ODATA_BASE}/dataset"
    )

    suffix = ".ics" if fmt in ICAL_FORMATS else (".xlsx" if fmt in SPREADSHEET_FORMATS else ".csv")
    fd, tmp = tempfile.mkstemp(suffix=suffix, prefix="odata-import-")
    os.close(fd)
    try:
        if fmt == "CSV" and datastore_active:
            # Clean, CKAN-normalised CSV — stream the datastore dump.
            await _download_dump(resource_id, tmp)
            loaded = await _load_csv_into(tmp, table)
        elif supported:
            if not url:
                raise ValueError("למשאב אין קובץ להורדה")
            await _download_file(url, tmp)
            # Parsing is CPU-bound (openpyxl / ical) — keep it off the event loop.
            columns, batches = await asyncio.to_thread(_open_for_load, tmp, fmt)
            if not columns:
                raise ValueError("לא נמצאו עמודות בקובץ")
            loaded = await _load_rows(table, columns, batches)
        else:
            # datastore-active but blank/unknown format → CSV dump.
            await _download_dump(resource_id, tmp)
            loaded = await _load_csv_into(tmp, table)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

    await _record_import(
        resource_id=resource_id, table=table, package_id=res.get("package_id"),
        dataset_name=dataset_name, title=title, organization=org, fmt=fmt,
        source_url=source_url, file_url=url or None,
        rows=loaded["rows"], columns=loaded["columns"])
    logger.info("odata import: %s -> odata.%s (%d rows, %d cols)",
                resource_id, table, loaded["rows"], loaded["columns"])
    return {"resource_id": resource_id, "table": table, "title": title,
            "organization": org, "rows": loaded["rows"],
            "columns": loaded["columns"], "source_url": source_url}


# ── background import jobs ───────────────────────────────────────────────────
# Parsing and loading a big spreadsheet takes minutes — the 548k-row "רשימת
# כתובות בישראל עם קואורדינטות" needs ~3 — which is far longer than the gateway
# will hold a request open, so an inline import returned a 504 no matter how
# well it went. The upload endpoint hands the file to one of these jobs and
# returns at once; the admin UI polls for the outcome. Single-process app (one
# uvicorn worker, see render.yaml), so an in-memory registry is enough.

_JOBS: dict[str, dict] = {}
_JOB_TASKS: set = set()
_JOB_TTL = 3600.0  # a finished job stays pollable for an hour

_JOB_PUBLIC = ("id", "resource_id", "title", "state", "rows", "columns",
               "table", "error")


def _prune_jobs() -> None:
    cutoff = time.time() - _JOB_TTL
    for jid in [j for j, v in _JOBS.items()
                if v.get("finished_at") and v["finished_at"] < cutoff]:
        _JOBS.pop(jid, None)


def job_status(job_id: str) -> dict | None:
    """Public view of an import job, or None if unknown/expired."""
    job = _JOBS.get(job_id)
    if not job:
        return None
    out = {k: job.get(k) for k in _JOB_PUBLIC}
    out["elapsed"] = round((job.get("finished_at") or time.time())
                           - job["started_at"], 1)
    return out


def start_upload_job(*, resource_id: str, fmt: str, dataset_name: str | None,
                     title: str | None, organization: str | None,
                     source_url: str | None, file_url: str | None,
                     tmp_path: str, filename: str | None = None) -> dict:
    """Run ``import_uploaded`` in the background; returns the job's first status.

    Owns the uploaded temp file: it is removed when the job ends, however it
    ends."""
    _prune_jobs()
    job_id = uuid.uuid4().hex
    job = {
        "id": job_id, "resource_id": resource_id,
        "title": (title or "").strip() or resource_id,
        "state": "running", "rows": 0, "columns": None, "table": None,
        "error": None, "started_at": time.time(), "finished_at": None,
    }
    _JOBS[job_id] = job

    async def _run() -> None:
        try:
            res = await import_uploaded(
                resource_id=resource_id, fmt=fmt, dataset_name=dataset_name,
                title=title, organization=organization, source_url=source_url,
                file_url=file_url, tmp_path=tmp_path, filename=filename,
                progress=lambda n: job.__setitem__("rows", n),
            )
            job.update(state="done", table=res["table"], rows=res["rows"],
                       columns=res["columns"], title=res["title"])
        except Exception as e:  # noqa: BLE001 — the job reports its own failure
            logger.exception("odata import job %s failed (%s)", job_id, resource_id)
            job.update(state="error", error=str(e) or e.__class__.__name__)
        finally:
            job["finished_at"] = time.time()
            try:
                os.remove(tmp_path)
            except OSError:
                pass

    task = asyncio.create_task(_run())
    _JOB_TASKS.add(task)                      # keep a ref so it isn't GC'd
    task.add_done_callback(_JOB_TASKS.discard)
    return job_status(job_id)


async def list_imports() -> list[dict]:
    """Registry rows whose table physically exists, newest first."""
    if not append_store.is_configured():
        return []
    pool = await append_store.get_pool()
    try:
        async with pool.acquire() as conn:
            await ensure_registry(conn)
            rows = await conn.fetch(f"""
                SELECT r.resource_id, r.table_name, r.dataset_name, r.title,
                       r.organization, r.format, r.source_url, r.source_file_url,
                       r.rows, r.columns, r.imported_at
                FROM {_qt(REGISTRY)} r
                JOIN pg_class c ON c.relname = r.table_name
                JOIN pg_namespace n ON n.oid = c.relnamespace
                                   AND n.nspname = '{SCHEMA}'
                ORDER BY r.imported_at DESC
            """)
    except Exception:  # noqa: BLE001 — schema not created yet
        logger.debug("odata: list_imports before first import", exc_info=True)
        return []
    return [{
        "resource_id": r["resource_id"], "table": r["table_name"],
        "dataset_name": r["dataset_name"], "title": r["title"],
        "organization": r["organization"], "format": r["format"],
        "source_url": r["source_url"], "source_file_url": r["source_file_url"],
        "rows": r["rows"], "columns": r["columns"], "imported_at": r["imported_at"],
    } for r in rows]


async def delete_import(resource_id: str) -> bool:
    """Drop an imported table and forget it. False if it wasn't imported."""
    if not append_store.is_configured():
        return False
    pool = await append_store.get_pool()
    async with pool.acquire() as conn:
        await ensure_registry(conn)
        row = await conn.fetchrow(
            f"SELECT table_name FROM {_qt(REGISTRY)} WHERE resource_id = $1",
            resource_id)
        if not row:
            return False
        await conn.execute(f"DROP TABLE IF EXISTS {_qt(row['table_name'])}")
        await conn.execute(
            f"DELETE FROM {_qt(REGISTRY)} WHERE resource_id = $1", resource_id)
    from app.services.data_catalog import invalidate_catalog_cache
    invalidate_catalog_cache()
    logger.info("odata import: dropped %s (%s)", row["table_name"], resource_id)
    return True
